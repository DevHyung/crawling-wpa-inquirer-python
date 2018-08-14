from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os

driver = webdriver.Chrome('./chromedriver')
driver.maximize_window()
driver.get('https://www.fangraphs.com/tools/wpa-inquirer')

def click_by_xpath(_xpath):
    tmpIdx = 1
    while True:
        if tmpIdx == 20:
            break
        else:
            tmpIdx += 1
        try:
            driver.find_element_by_xpath(_xpath).click()
            break
        except:
            time.sleep(0.5)
def save_excel(_filename,_datalist):
    FILENAME = './'+_filename
    if os.path.isfile(FILENAME):  # 파일있는 경우
        wb = load_workbook(filename=FILENAME)
        sheet1 = wb[wb.sheetnames[0]]
        sheet1.append(_datalist)
        wb.save(FILENAME)
    else:  # 파일 없는 경우
        # 엑셀파일 초기설정
        book = Workbook()
        # 시트 설정
        sheet1 = book.active
        sheet1.column_dimensions['A'].width = 10
        sheet1.column_dimensions['B'].width = 10
        sheet1.column_dimensions['C'].width = 10
        sheet1.column_dimensions['D'].width = 10

        # 저장
        sheet1.append(['RunEnvironment', 'BaseSituation', 'Inning', 'Outs', 'Minus10', 'Minus9', 'Minus8', 'Minus7', 'Minus6', 'Minus5', 'Minus4', 'Minus3', 'Minus2', 'Minus1', 'Zero', 'Plus1', 'Plus2', 'Plus3', 'Plus4', 'Plus5', 'Plus6', 'Plus7', 'Plus8', 'Plus9', 'Plus10'])
        sheet1.append(_datalist)
        book.save(FILENAME)

DELAY = 3
# 원래 1,9까지 사실상 1~8까지
# 진행중
# 2,3,4,5, 6
# 완료
# 1
for i in range(8,9): # 8개
    # 첫번째
    HOMEFILE = 'home{}.xlsx'.format(i)
    LEVERFILE = 'leverage{}.xlsx'.format(i)
    print(">>>", HOMEFILE, '시작')

    driver.find_element_by_xpath('//*[@id="rcbRun_Input"]').click()
    click_by_xpath('//*[@id="rcbRun_DropDown"]/div/ul/li[{}]'.format(i))
    #driver.find_element_by_xpath('//*[@id="rcbRun_DropDown"]/div/ul/li[{}]'.format(i)).click() # 첫번쨰
    #time.sleep(DELAY)

    # 두번쨰
    for twoIdx in range(1,9): #8개
        #driver.find_element_by_xpath('//*[@id="rcbBase_Input"]').click()
        click_by_xpath('//*[@id="rcbBase_Input"]')
        click_by_xpath('//*[@id="rcbBase_DropDown"]/div/ul/li[{}]'.format(twoIdx))
        #time.sleep(DELAY)

        # 세번째
        for threeIdx in range(1, 19):  # 8개
            #driver.find_element_by_xpath('//*[@id="rcbInning_Input"]').click()
            click_by_xpath('//*[@id="rcbInning_Input"]')
            click_by_xpath('//*[@id="rcbInning_DropDown"]/div/ul/li[{}]'.format(threeIdx))
            #time.sleep(DELAY)

            # 네번째 rcbOuts_Input
            for fourIdx in range(1, 4):  # 3개
                #driver.find_element_by_xpath('//*[@id="rcbOuts_Input"]').click()
                click_by_xpath('//*[@id="rcbOuts_Input"]')
                click_by_xpath('//*[@id="rcbOuts_DropDown"]/div/ul/li[{}]'.format(fourIdx))
                #time.sleep(DELAY)
                homedatalist = []
                leveragedatalist = []
                # 다섯번째 rcbScore_Input
                for fiveIdx in range(1, 22):  # 21개
                    #driver.find_element_by_xpath('//*[@id="rcbScore_Input"]').click()
                    click_by_xpath('//*[@id="rcbScore_Input"]')
                    click_by_xpath(
                        '//*[@id="rcbScore_DropDown"]/div/ul/li[{}]'.format(fiveIdx))

                    click_by_xpath('//*[@id="rcbScore_Input"]')
                    click_by_xpath('//*[@id="rcbScore_Input"]')
                    time.sleep(0.1)
                    #=== TEST PRINT
                    bs4 = BeautifulSoup(driver.page_source, 'lxml')
                    runEnv = bs4.find('input', id='rcbRun_Input')['value']
                    baseSitu = bs4.find('input',id='rcbBase_Input')['value']
                    inning = bs4.find('input', id='rcbInning_Input')['value']
                    outs = bs4.find('input', id='rcbOuts_Input')['value']
                    runDifferential = bs4.find('input', id='rcbScore_Input')['value']
                    homeExp = driver.find_element_by_xpath('//*[@id="radAjax"]/table/tbody/tr[6]/td[2]').text.strip()
                    Leverage =driver.find_element_by_xpath('//*[@id="radAjax"]/table/tbody/tr[8]/td[2]').text.strip()
                    if len(homedatalist) == 0:
                        homedatalist.extend([runEnv,baseSitu,inning,outs,homeExp])
                        print(runEnv, baseSitu, inning, outs, runDifferential, homeExp, Leverage)
                    else:
                        homedatalist.append(homeExp)
                    if len(leveragedatalist) == 0:
                        leveragedatalist.extend([runEnv,baseSitu,inning,outs,Leverage])
                    else:
                        leveragedatalist.append(Leverage)
                #print(homedatalist)
                #print(leveragedatalist)
                save_excel(HOMEFILE, homedatalist)
                save_excel(LEVERFILE, leveragedatalist)
driver.quit()
